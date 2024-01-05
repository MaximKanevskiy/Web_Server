import openpyxl
import json
import re


class DayOfWeek:
    def __init__(self, pair_type, subject_name, classroom, lecturer):
        self.тип_пары = pair_type
        self.предмет = subject_name
        self.аудитория = classroom
        self.лектор = lecturer
        self.сообщение = ""

    def to_dict(self):
        return {
            "тип пары": self.тип_пары,
            "предмет": self.предмет,
            "аудитория": self.аудитория,
            "лектор": self.лектор,
            "сообщение": self.сообщение
        }


def main():
    wb = openpyxl.load_workbook('schedule.xlsx')
    ws = wb.active

    week = []

    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    start_rows = [5, 40, 70, 105, 135]
    end_rows = [39, 69, 104, 134, 164]
    pair_names = ["первая", "вторая", "третья", "четвертая", "пятая", "шестая", "седьмая"]

    for day_index, (day, start, end) in enumerate(zip(days, start_rows, end_rows)):
        day_schedule = {"SchID": f"ИВТ-231-1_нечет_{day_index + 1}"}
        for pair_index, row in enumerate(range(start, end + 1, 5)):
            if ws.cell(row=row, column=2).value is not None:
                pair_number = pair_names[pair_index]
                pair_type = ws.cell(row=row, column=3).value
                subject_name = ws.cell(row=row, column=4).value
                lecturer = ws.cell(row=row + 1, column=4).value
                classroom = ws.cell(row=row + 2, column=4).value

                # Преобразование типа пары
                if pair_type == "ЛК":
                    pair_type = "лекция"
                elif pair_type == "ПЗ":
                    pair_type = "практика"

                # Извлечение номера аудитории
                if classroom is not None:
                    match = re.search(r'\d+[А-Яа-я]', classroom)
                    classroom = match.group() if match else ""
                else:
                    classroom = ""

                # Преобразование названия предмета в нижний регистр
                if subject_name is not None:
                    subject_name = subject_name.lower()

                if pair_type in ["", None] and subject_name in ["", None] and lecturer in ["", None] and classroom in [
                    "", None]:
                    day_schedule[f"{pair_number} пара"] = "нет"
                else:
                    day_of_week = DayOfWeek(pair_type, subject_name, classroom, lecturer)
                    day_schedule[f"{pair_number} пара"] = day_of_week.to_dict()
        week.append(day_schedule)

    with open('нечет_ивт-231-1.json', 'w', encoding='utf-8') as f:
        json.dump(week, f, ensure_ascii=False, indent=4)


def main2():
    wb = openpyxl.load_workbook('schedule.xlsx')
    ws = wb.active

    week = []

    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    start_rows = [5, 40, 70, 105, 135]
    end_rows = [39, 69, 104, 134, 164]
    pair_names = ["первая", "вторая", "третья", "четвертая", "пятая", "шестая", "седьмая"]

    for day_index, (day, start, end) in enumerate(zip(days, start_rows, end_rows)):
        day_schedule = {"SchID": f"ИВТ-231-2_нечет_{day_index+1}"}
        for pair_index, row in enumerate(range(start, end + 1, 5)):
            if ws.cell(row=row, column=2).value is not None:
                pair_number = pair_names[pair_index]
                pair_type = ws.cell(row=row, column=3).value
                subject_name = ws.cell(row=row, column=5).value
                subject_name = subject_name if subject_name is None else subject_name.lower()
                for range_ in ws.merged_cells.ranges:
                    if f'E{row}' in range_:
                        subject_name = ws.cell(row=range_.min_row, column=4).value
                        subject_name = subject_name if subject_name is None else subject_name.lower()
                        lecturer = ws.cell(row=row + 1, column=4).value
                        classroom = ws.cell(row=row + 2, column=4).value
                        break
                else:
                    lecturer = ws.cell(row=row + 1, column=5).value
                    classroom = ws.cell(row=row + 2, column=5).value

                # Преобразование типа пары
                if pair_type == "ЛК":
                    pair_type = "лекция"
                elif pair_type == "ПЗ":
                    pair_type = "практика"

                # Извлечение номера аудитории
                if classroom is not None and subject_name != "Физическая культура":
                    match = re.search(r'\d+[А-Яа-я]', classroom)
                    classroom = match.group() if match else ""
                elif classroom is not None and subject_name == "Физическая культура":
                    classroom = ws.cell(row=row + 2, column=5).value
                else:
                    classroom = ""

                if pair_type in ["", None] and subject_name in ["", None] and lecturer in ["", None] and classroom in ["", None]:
                    day_schedule[f"{pair_number} пара"] = "нет"
                else:
                    day_of_week = DayOfWeek(pair_type, subject_name, classroom, lecturer)
                    day_schedule[f"{pair_number} пара"] = day_of_week.to_dict()
        week.append(day_schedule)

    with open('нечет_ивт-231-2.json', 'w', encoding='utf-8') as f:
        json.dump(week, f, ensure_ascii=False, indent=4)


if __name__ == "__main__":
    main()
    main2()
